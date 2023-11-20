import openpyxl
import csv

class Course:
    def __init__(self, ccode, cname, cdate):
        self.ccode = ccode
        self.cname = cname
        self.cdates = cdate
        self.sections = []

    def get_all_sections(self):
        return [section.sid for section in self.sections]

    def __str__(self):
        return f"Course Code: {self.ccode}, Course Name: {self.cname}, Exam Dates: {', '.join(self.cdates)}"

    def populate_sections(self, sid, day, slot):
        existing_section = next((section for section in self.sections if section.sid == sid), None)

        if existing_section:
            existing_section.add_schedule(day, slot)
        else:
            snew = Sections(sid, "lecture")  
            snew.add_schedule(day, slot)
            self.sections.append(snew)




class Sections:
    def __init__(self, sid, stype, schedule=None):
        self.sid = sid
        self.stype = stype
        self.schedule = {} if schedule is None else schedule

    def add_schedule(self, day, slot):
        if day not in self.schedule:
            self.schedule[day] = set()  
        self.schedule[day].add(slot)

    def has_clash(self, other_section):
        if not isinstance(other_section, Sections):
            return False

        for day, slots in self.schedule.items():
            if day in other_section.schedule:
                if any(slot in other_section.schedule[day] for slot in slots):
                    return True
        return False

    def __str__(self):
        sinfo = ", ".join([f"{day}: {', '.join(slots)}" for day, slots in self.schedule.items()])
        return f"Section ID: {self.sid}, Type: {self.stype}, Schedule: {sinfo}\n"

    

class Timetable:
    def __init__(self):
        self.courses = []

    def enroll_subject(self, course):
        self.courses.append(course)

    def check_clashes(self):
        for i, course1 in enumerate(self.courses[:-1]):
            for course2 in self.courses[i + 1:]:
                for section1 in course1.sections:
                    for section2 in course2.sections:
                        if section1.has_clash(section2):
                            print(f"Clash detected between {course1.ccode} ({section1.sid}) and {course2.ccode} ({section2.sid}).")

    def export_to_csv(self, filename):
        with open(filename, 'w', newline='') as csvfile:
            fieldnames = ['Course Code', 'Course Name', 'Exam Dates', 'Sections']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()
            for course in self.courses:
                writer.writerow({
                    'Course Code': course.ccode,
                    'Course Name': course.cname,
                    'Exam Dates': ', '.join(course.cdates),
                    'Sections': ', '.join(course.get_all_sections())
                })

    def __str__(self):
        return "\n".join(str(course) for course in self.courses)

    def populate_course(self, filename):
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        for row in range(2, sheet.max_row + 1):
            ccode = sheet.cell(row=row, column=1).value  
            cname = sheet.cell(row=row, column=2).value  
            cdate = sheet.cell(row=row, column=3).value.split(', ') if sheet.cell(row=row, column=3).value else []
            cnew = Course(ccode, cname, cdate)
            self.enroll_subject(cnew)

        wb.close()

if __name__ == "__main__":
    timetable = Timetable()

    course1 = Course("F112", "Thermodynamics", ["11/10/2023", "11/12/2023"])
    course1.populate_sections("L1", "Tuesday", "9:00 AM - 10:00 AM")
    course1.populate_sections("L2", "Thursday", "9:00 AM - 10:00 AM")
    course1.populate_sections("L3", "Saturday", "9:00 AM - 10:00 AM")
    course1.populate_sections("T1", "Monday", "8:00 AM - 9:00 AM")

    course2 = Course("MATH F111", "MATHEMATICS-1", ["13/10/2023", "18/12/2023"])
    course2.populate_sections("L1", "Tuesday", "9:00 AM - 10:00 AM")
    course2.populate_sections("L2", "Wednesday", "12:00 PM - 1:00 PM")
    course2.populate_sections("L3", "Friday", "12:00 PM - 1:00 PM")
    course2.populate_sections("T1", "Thursday", "8:00 AM - 9:00 AM")

    timetable.enroll_subject(course1)
    timetable.enroll_subject(course2)

    timetable.check_clashes()
    timetable.export_to_csv("timetable.csv")
    print(timetable)



if __name__ == "__main__":
    timetable = Timetable()
    timetable.populate_course("courses.xlsx")
    print(timetable)

