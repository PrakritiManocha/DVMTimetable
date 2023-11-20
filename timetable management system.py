import openpyxl
import csv
from collections import defaultdict

class Course:
    def __init__(self, ccode, cname, cdate):
        self.ccode = ccode
        self.cname = cname
        self.cdates = cdate
        self.sections = []

    def get_all_sections(self):
        return [section.sid for section in self.sections]

    def __str__(self):
        return f"Course Code: {self.ccode}, Course Name: {self.cname}, Exam Dates: {', '.join(self.cdates)}"

    def populate_sections(self, sid, day, slot):
        snew = Sections(sid, day, slot)
        self.sections.append(snew)



class Sections:
    def __init__(self, sid, stype, schedule=None):
        self.sid = sid
        self.stype = stype
        self.schedule = {} if schedule is None else schedule

    def add_schedule(self, day, slot):
        self.schedule[day] = slot

    def __str__(self):
        sinfo = ", ".join([f"{day}: {slot}" for day, slot in self.schedule.items()])
        return f"Section ID: {self.sid}, Type: {self.stype}, Schedule: {sinfo}\n"


# Example usage of Sections class & Course class
if __name__ == "__main__":
    section = Sections("L1", "lecture")
    section.add_schedule("Monday", "10:00 AM - 11:00 AM")
    section.add_schedule("Wednesday", "8:00 PM - 9:00 AM")
    print(section)

if __name__ == "__main__":
    course = Course("CSF111", "Computer Programming", ["25/11/23", "15/12/23"])
    course.populate_sections("L1", "Monday", "10:00 AM - 11:00 AM")
    course.populate_sections("L2", "Wednesday", "10:00 AM - 11:00 AM")
    course.populate_sections("L3", "Friday", "10:00 AM - 11:00 AM")
    course.populate_sections("P1", "Friday", "3:00 PM - 5:00 PM")

    print(course)
    print(course.get_all_sections())

class Timetable:
    def __init__(self):
        self.courses = []

    def enroll_subject(self, course):
        self.courses.append(course)

    def check_clashes(self, day, slot):
        for course in self.courses:
            for section in course.sections:
                if day in section.schedule and slot == section.schedule[day]:
                    print(f"Clash detected for {course.ccode} on {day} at {slot} with {section.sid}.")

    def export_to_csv(self, filename):
        with open(filename, 'w', newline='') as csvfile:
            fieldnames = ['Course Code', 'Course Name', 'Exam Dates', 'Sections']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()
            for course in self.courses:
                writer.writerow({
                    'Course Code': course.ccode,
                    'Course Name': course.cname,
                    'Exam Dates': ', '.join(course.cdates),
                    'Sections': ', '.join(course.get_all_sections())
                })

    def __str__(self):
        return "\n".join(str(course) for course in self.courses)

    def populate_course(self, filename):
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        for row in range(2, sheet.max_row + 1):
            ccode = sheet.cell(row=row, column=1).value  
            cname = sheet.cell(row=row, column=2).value  
            cdate = sheet.cell(row=row, column=3).value.split(', ') if sheet.cell(row=row, column=3).value else []
            cnew = Course(ccode, cname, cdate)
            self.enroll_subject(cnew)

        wb.close()



if __name__ == "__main__":

    timetable = Timetable()

    course1 = Course("F112", "Thermodynamics", ["11/10/2023", "11/12/2023"])
    course1.populate_sections("L1", "Tuesday", "9:00 AM - 10:00 AM")
    course1.populate_sections("L2", "Thursday", "9:00 AM - 10:00 AM")
    course1.populate_sections("L3", "Saturday", "9:00 AM - 10:00 AM")
    course1.populate_sections("T1", "Monday", "8:00 AM - 9:00 AM")

    course2 = Course("MATH F111", "MATHEMATICS-1", ["13/10/2023", "18/12/2023"])
    course2.populate_sections("L1", "Monday", "12:00 PM - 1:00 PM")
    course2.populate_sections("L2", "Wednesday", "12:00 PM - 1:00 PM")
    course2.populate_sections("L3", "Friday", "12:00 PM - 1:00 PM")
    course2.populate_sections("T1", "Thursday", "8:00 AM - 9:00 AM")

    timetable.enroll_subject(course1)
    timetable.enroll_subject(course2)


    timetable.export_to_csv("timetable.csv")

    print(timetable)




if __name__ == "__main__":
    timetable = Timetable()

  
    timetable.populate_course("courses.xlsx")


if __name__ == "__main__":

    techre = Course("F111", "Technical Report Writing", ["7/8/23, 15/12/23"])
    techre.populate_sections("L1", "Tuesday", "09:00 AM - 11:00 AM")
    timetable.enroll_subject(techre)




print(timetable)


